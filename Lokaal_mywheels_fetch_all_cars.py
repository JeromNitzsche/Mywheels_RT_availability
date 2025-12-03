import requests
import openpyxl
from openpyxl import Workbook
import time
import os

# ==========================================
# INSTELLINGEN
# ==========================================

START_ID = 28600
END_ID   = 28799

# Excel bestand in dezelfde map
EXCEL_NAME = "Source_Car_ID_local.xlsx"
EXCEL_PATH = EXCEL_NAME   # <— hier GEEN os.path.dirname(__file__)

API_URL = "https://prod-api.mywheels.nl/api/"

# Verwachte kolomtitels
HEADERS = [
    "id", "registration_plate", "brand", "model", "Aantal", "city",
    "street", "latitude", "longitude", "fuelType", "color", "model_full"
]


# ==========================================
# API FUNCTIE
# ==========================================

def fetch_car(resource_id):

    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
    }

    payload = [{
        "jsonrpc": "2.0",
        "id": 1,
        "method": "resource.get",
        "params": {
            "resource": resource_id
        }
    }]

    try:
        r = requests.post(API_URL, json=payload, headers=headers, timeout=10)

        try:
            data = r.json()
        except:
            print(f"⚠️ Geen JSON voor ID {resource_id} → HTTP {r.status_code}")
            return None

        if not isinstance(data, list) or "result" not in data[0]:
            return None

        car = data[0]["result"]

        if not car.get("registrationPlate"):
            return None

        return {
            "id": resource_id,
            "plate": car.get("registrationPlate"),
            "brand": car.get("brand"),
            "model": car.get("model"),
            "city": car.get("city"),
            "street": car.get("location"),
            "latitude": car.get("latitude"),
            "longitude": car.get("longitude"),
            "fuelType": car.get("fuelType"),
            "color": car.get("color"),
            "model_full": car.get("model")
        }

    except Exception as e:
        print(f"⚠️ API-fout bij ID {resource_id} → {e}")
        return None


# ==========================================
# EXCEL LADEN OF CREËREN
# ==========================================

def load_or_create_excel(path):
    if not os.path.exists(path):
        print("📄 Excel bestaat niet → nieuwe maken...")
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        return wb, ws

    print("📁 Excel gevonden → openen...")
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    first_row = [cell.value for cell in ws[1]]
    if first_row != HEADERS:
        print("⚠️ Headers ongeldig → resetten...")
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADERS)

    return wb, ws


# ==========================================
# HOOFDLOGICA
# ==========================================

print("📄 Excel laden...")
wb, ws = load_or_create_excel(EXCEL_PATH)

index = {}

for row in ws.iter_rows(min_row=2):
    rid = row[0].value
    if isinstance(rid, int):
        index[rid] = row

print("🚗 Start ophalen voertuigen...")

updated = 0
inserted = 0

for rid in range(START_ID, END_ID + 1):
    car = fetch_car(rid)

    if not car or not car["plate"]:
        continue

    if rid in index:
        row = index[rid]
        row[1].value = car["plate"]
        row[2].value = car["brand"]
        row[3].value = car["model"]
        row[4].value = 1
        row[5].value = car["city"]
        row[6].value = car["street"]
        row[7].value = car["latitude"]
        row[8].value = car["longitude"]
        row[9].value = car["fuelType"]
        row[10].value = car["color"]
        row[11].value = car["model_full"]
        updated += 1
    else:
        ws.append([
            car["id"], car["plate"], car["brand"], car["model"], 1,
            car["city"], car["street"], car["latitude"], car["longitude"],
            car["fuelType"], car["color"], car["model_full"]
        ])
        inserted += 1

    time.sleep(0.25)

print("📊 Sorteren op city...")

rows = list(ws.iter_rows(values_only=True))
header = rows[0]
entries = rows[1:]

entries_sorted = sorted(entries, key=lambda x: (x[5] or ""))

ws.delete_rows(1, ws.max_row)
ws.append(header)
for row in entries_sorted:
    ws.append(row)

print("💾 Opslaan...")
wb.save(EXCEL_PATH)

print(f"🎉 KLAAR! {updated} bijgewerkt, {inserted} toegevoegd!")
print(f"📁 Bestand opgeslagen als: {EXCEL_PATH}")
